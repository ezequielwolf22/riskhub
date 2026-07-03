"""Servicio de importación/exportación Excel para el módulo BCP."""
import html
import io
import logging

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.i18n import t as _t

logger = logging.getLogger("riskhub.bcp_excel")

_HEADER_FILL = PatternFill("solid", fgColor="59008D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(wrap_text=True, horizontal="center", vertical="center")


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    ws.row_dimensions[1].height = 30


def generate_excel_template(lang: str = "es") -> bytes:
    """Genera la plantilla Excel BCP para importación."""
    wb = openpyxl.Workbook()

    # Hoja 1: Procesos BIA
    # NOTA: los nombres de hoja ("Procesos", "Dependencias", "Proveedores BCM")
    # son identificadores del parser (parse_excel_preview) y NO se traducen.
    ws_proc = wb.active
    ws_proc.title = "Procesos"
    _write_header(ws_proc, _t("bcp_excel.tpl_proc_headers", lang).split("|"))
    yes = _t("bcp_excel.yes_value", lang)
    ex_erp = _t("bcp_excel.ex_proc_erp", lang)
    ex_portal = _t("bcp_excel.ex_proc_portal", lang)
    ws_proc.append([ex_erp, "critical", 4, 1, 8,
                    _t("bcp_excel.ex_proc_erp_desc", lang), "admin@empresa.com", 1])
    ws_proc.append([ex_portal, "high", 8, 4, 24,
                    _t("bcp_excel.ex_proc_portal_desc", lang), "", 2])
    ws_proc.append([_t("bcp_excel.ex_proc_backup", lang), "medium", 24, 4, 48,
                    _t("bcp_excel.ex_proc_backup_desc", lang), "", 3])

    # Hoja 2: Dependencias
    ws_dep = wb.create_sheet("Dependencias")
    _write_header(ws_dep, _t("bcp_excel.tpl_dep_headers", lang).split("|"))
    ws_dep.append([ex_erp, "IT_system", _t("bcp_excel.ex_dep_db", lang), 2, yes])
    ws_dep.append([ex_erp, "personnel", _t("bcp_excel.ex_dep_dba", lang), 4, yes])
    ws_dep.append([ex_portal, "IT_system", _t("bcp_excel.ex_dep_web", lang), 4, yes])

    # Hoja 3: Proveedores BCM
    ws_sup = wb.create_sheet("Proveedores BCM")
    _write_header(ws_sup, _t("bcp_excel.tpl_sup_headers", lang).split("|"))
    ws_sup.append(["AWS", "critical", 1])
    ws_sup.append(["Microsoft 365", "high", 4])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_excel_preview(content: bytes) -> dict:
    """Parsea el Excel BCP y devuelve preview con errores."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    errors = []
    processes = []
    dependencies = []
    suppliers = []

    # -- Procesos --
    if "Procesos" in wb.sheetnames:
        ws = wb["Procesos"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name:
                continue
            crit = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "medium"
            if crit not in ("critical", "high", "medium", "low"):
                # html.escape previene XSS cuando el error se renderiza en el frontend
                errors.append(f"Procesos fila {i}: criticidad inválida '{html.escape(crit)}'")
                crit = "medium"
            processes.append({
                "name": name,
                "criticality": crit,
                "rto_hours": _safe_int(row[2]) if len(row) > 2 else None,
                "rpo_hours": _safe_int(row[3]) if len(row) > 3 else None,
                "mtpd_hours": _safe_int(row[4]) if len(row) > 4 else None,
                "description": str(row[5]).strip() if len(row) > 5 and row[5] else None,
                "owner_email": str(row[6]).strip() if len(row) > 6 and row[6] else None,
                "priority": _safe_int(row[7]) if len(row) > 7 else None,
            })
    else:
        errors.append("Hoja 'Procesos' no encontrada en el archivo")

    # -- Dependencias --
    if "Dependencias" in wb.sheetnames:
        ws = wb["Dependencias"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            dependencies.append({
                "process_name": str(row[0]).strip() if row[0] else "",
                "dependency_type": str(row[1]).strip() if len(row) > 1 and row[1] else "IT_system",
                "name": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "rto_hours": _safe_int(row[3]) if len(row) > 3 else None,
                "is_critical": str(row[4]).strip().lower() in ("si", "yes", "true", "1")
                    if len(row) > 4 and row[4] else False,
            })

    # -- Proveedores BCM --
    if "Proveedores BCM" in wb.sheetnames:
        ws = wb["Proveedores BCM"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            crit = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "medium"
            if crit not in ("critical", "high", "medium", "low"):
                crit = "medium"
            suppliers.append({
                "supplier_name": str(row[0]).strip(),
                "criticality": crit,
                "rto_impact_hours": _safe_int(row[2]) if len(row) > 2 else None,
            })

    return {
        "errors": errors,
        "summary": {
            "processes_found": len(processes),
            "dependencies_found": len(dependencies),
            "suppliers_found": len(suppliers),
        },
        "processes": processes[:3],   # solo primeras 3 filas para preview UI
        "all_processes": processes,
        "all_dependencies": dependencies,
        "all_suppliers": suppliers,
    }


def confirm_excel_import(db: Session, preview: dict, org_id: int) -> dict:
    """Crea los registros en BD a partir del preview parseado."""
    from app.models import BusinessProcess, BCPDependency, BCPSupplierLink, Supplier, User

    created = {"processes": 0, "dependencies": 0, "supplier_links": 0}
    proc_map: dict[str, int] = {}  # nombre → id

    # Crear procesos
    for pd in preview.get("all_processes", []):
        name = pd["name"]
        existing = db.query(BusinessProcess).filter_by(
            organization_id=org_id, name=name
        ).first()
        if existing:
            proc_map[name] = existing.id
            continue
        owner = None
        if pd.get("owner_email"):
            owner = db.query(User).filter_by(email=pd["owner_email"]).first()
        p = BusinessProcess(
            organization_id=org_id,
            name=name,
            criticality=pd.get("criticality", "medium"),
            rto_hours=pd.get("rto_hours"),
            rpo_hours=pd.get("rpo_hours"),
            mtpd_hours=pd.get("mtpd_hours"),
            description=pd.get("description"),
            priority=pd.get("priority"),
            owner_id=owner.id if owner else None,
        )
        db.add(p)
        db.flush()
        proc_map[name] = p.id
        created["processes"] += 1

    # Crear dependencias
    for dd in preview.get("all_dependencies", []):
        proc_id = proc_map.get(dd.get("process_name", ""))
        if not proc_id or not dd.get("name"):
            continue
        dep = BCPDependency(
            organization_id=org_id,
            process_id=proc_id,
            dependency_type=dd.get("dependency_type", "IT_system"),
            name=dd["name"],
            rto_hours=dd.get("rto_hours"),
            is_critical=dd.get("is_critical", False),
        )
        db.add(dep)
        created["dependencies"] += 1

    # Crear vínculos de proveedor BCM
    for sd in preview.get("all_suppliers", []):
        sup = db.query(Supplier).filter_by(
            organization_id=org_id, name=sd["supplier_name"]
        ).first()
        if not sup:
            continue
        already = db.query(BCPSupplierLink).filter_by(
            organization_id=org_id, supplier_id=sup.id
        ).first()
        if already:
            continue
        sl = BCPSupplierLink(
            organization_id=org_id,
            supplier_id=sup.id,
            criticality=sd.get("criticality", "medium"),
            rto_impact_hours=sd.get("rto_impact_hours"),
        )
        db.add(sl)
        created["supplier_links"] += 1

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("Excel import rollback: %s", exc)
        raise ValueError(f"Error guardando datos: {exc}") from exc

    return created


def _safe_int(val) -> int | None:
    try:
        return int(float(str(val))) if val is not None else None
    except (ValueError, TypeError):
        return None


def export_org_data(db: Session, org_id: int) -> bytes:
    """Exporta todos los datos BCP de la organización como Excel."""
    from app.models import (
        BusinessProcess, BCPDependency, BCPPlan, BCPTest, BCPSupplierLink, Supplier,
    )
    from app.services.bcp_service import bia_completeness

    wb = openpyxl.Workbook()

    # Hoja 1: Procesos
    ws_proc = wb.active
    ws_proc.title = "Procesos"
    _write_header(ws_proc, [
        "ID", "Nombre", "Criticidad", "Prioridad", "RTO (h)", "RPO (h)", "MTPD (h)",
        "MBCO", "Impacto Financiero", "Impacto Reputacional", "Impacto Legal",
        "Impacto Operacional", "Staff min recuperacion", "Criterios activacion",
        "Procedimiento alternativo", "BIA %", "Ultimo test", "Resultado test",
    ])
    procs = db.query(BusinessProcess).filter_by(organization_id=org_id).all()
    for p in procs:
        try:
            bia = bia_completeness(None, p)
        except Exception:
            bia = {"pct": 0}
        ws_proc.append([
            p.id, p.name, p.criticality, p.priority,
            p.rto_hours, p.rpo_hours, p.mtpd_hours,
            p.mbco, p.financial_impact, p.reputational_impact, p.legal_impact,
            p.operational_impact, p.min_recovery_staff,
            p.activation_criteria, p.alternative_procedure,
            bia["pct"],
            p.last_tested_at.strftime("%Y-%m-%d") if p.last_tested_at else "",
            p.test_result or "",
        ])

    # Hoja 2: Dependencias
    ws_dep = wb.create_sheet("Dependencias")
    _write_header(ws_dep, [
        "ID", "Proceso", "Tipo", "Nombre", "Qty normal", "Qty recuperacion",
        "RTO necesario (h)", "Es critico", "Alternativa",
    ])
    deps = db.query(BCPDependency).filter_by(organization_id=org_id).all()
    proc_map = {p.id: p.name for p in procs}
    for d in deps:
        ws_dep.append([
            d.id, proc_map.get(d.process_id, d.process_id),
            d.dependency_type, d.name, d.qty_normal, d.qty_recovery,
            d.rto_hours, "Si" if d.is_critical else "No", d.alternative or "",
        ])

    # Hoja 3: Planes
    ws_plan = wb.create_sheet("Planes")
    _write_header(ws_plan, [
        "ID", "Codigo", "Tipo", "Nombre", "Version", "Estado",
        "Alcance", "Criterios activacion", "Ultima prueba", "Revision",
    ])
    plans = db.query(BCPPlan).filter_by(organization_id=org_id).all()
    for p in plans:
        ws_plan.append([
            p.id, p.code, p.plan_type, p.name, p.version, p.status,
            (p.scope or "")[:200], (p.activation_criteria or "")[:200],
            p.last_exercised_at.strftime("%Y-%m-%d") if p.last_exercised_at else "",
            p.review_date.strftime("%Y-%m-%d") if p.review_date else "",
        ])

    # Hoja 4: Tests
    ws_test = wb.create_sheet("Tests")
    _write_header(ws_test, [
        "ID", "Codigo", "Tipo", "Programado", "Realizado", "Resultado",
        "Objetivo", "Hallazgos", "Lecciones aprendidas",
    ])
    tests = db.query(BCPTest).filter_by(organization_id=org_id).all()
    for t in tests:
        ws_test.append([
            t.id, t.code, t.test_type,
            t.scheduled_at.strftime("%Y-%m-%d") if t.scheduled_at else "",
            t.conducted_at.strftime("%Y-%m-%d") if t.conducted_at else "",
            t.result or "", t.objective or "", t.findings or "", t.lessons_learned or "",
        ])

    # Hoja 5: Proveedores BCM
    ws_sup = wb.create_sheet("Proveedores BCM")
    _write_header(ws_sup, [
        "ID", "Proveedor", "Criticidad BCM", "RTO impacto (h)",
        "Plan contingencia", "SLA contrato (h)", "Ultima revision",
    ])
    slinks = db.query(BCPSupplierLink).filter_by(organization_id=org_id).all()
    for sl in slinks:
        sup = db.get(Supplier, sl.supplier_id)
        ws_sup.append([
            sl.id, sup.name if sup else sl.supplier_id,
            sl.criticality, sl.rto_impact_hours,
            "Si" if sl.has_contingency_plan else "No",
            sl.contract_sla_hours,
            sl.last_review_date.strftime("%Y-%m-%d") if sl.last_review_date else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def ai_parse_any_format(content: bytes, filename: str, api_key: str,
                              model: str = "claude-haiku-4-5") -> dict:
    """Usa Claude para parsear cualquier Excel/CSV y mapear a campos BCP."""
    import anthropic
    import csv
    import json

    raw_rows = []
    headers = []

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(text.splitlines())
        rows = list(reader)
        if rows:
            headers = rows[0]
            raw_rows = rows[1:21]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if all_rows:
            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            raw_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in all_rows[1:21]
            ]

    if not headers:
        return {
            "errors": ["No se detectaron columnas en el archivo"],
            "summary": {"processes_found": 0, "dependencies_found": 0, "suppliers_found": 0},
            "processes": [], "all_processes": [], "all_dependencies": [], "all_suppliers": [],
            "ai_mapping": {}, "ai_warnings": [], "ai_confidence": 0,
        }

    rows_text = "\n".join(
        [", ".join(str(h) for h in headers)] +
        [", ".join(str(c) for c in row) for row in raw_rows[:5]]
    )

    prompt = f"""Analiza este archivo Excel/CSV que contiene datos de BCP (Business Continuity Planning).

Columnas detectadas: {headers}

Primeras filas de datos:
{rows_text}

Tu tarea:
1. Identifica que columna corresponde a cada campo BCP de la lista de abajo.
2. Lista los campos BCP OBLIGATORIOS que faltan en el archivo.
3. Da advertencias sobre datos que parecen incorrectos o incompletos.

Campos BCP disponibles para procesos:
- name (OBLIGATORIO): nombre del proceso
- criticality: critical/high/medium/low
- rto_hours: Recovery Time Objective en horas (numero)
- rpo_hours: Recovery Point Objective en horas (numero)
- mtpd_hours: Maximum Tolerable Period of Disruption en horas (numero)
- description: descripcion del proceso
- priority: prioridad numerica
- activation_criteria: criterios de activacion del plan
- owner_email: email del propietario

Responde UNICAMENTE con JSON valido en este formato exacto:
{{
  "mapping": {{
    "name": "nombre_columna_detectada_o_null",
    "criticality": "nombre_columna_detectada_o_null",
    "rto_hours": "nombre_columna_detectada_o_null",
    "rpo_hours": "nombre_columna_detectada_o_null",
    "mtpd_hours": "nombre_columna_detectada_o_null",
    "description": "nombre_columna_detectada_o_null",
    "priority": "nombre_columna_detectada_o_null",
    "activation_criteria": "nombre_columna_detectada_o_null",
    "owner_email": "nombre_columna_detectada_o_null"
  }},
  "missing_required": ["lista de campos obligatorios que faltan"],
  "warnings": ["avisos sobre datos que parecen incorrectos o incompletos"],
  "confidence": 0.85
}}"""

    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model=model,
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = message.content[0].text.strip()
    # Limpiar markdown code blocks si Claude los incluye
    if raw.startswith("```"):
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else raw
        if raw.startswith("json"):
            raw = raw[4:]
    ai_result = json.loads(raw.strip())

    mapping = ai_result.get("mapping", {})
    warnings = ai_result.get("warnings", [])
    missing = ai_result.get("missing_required", [])

    # Aplicar el mapping para extraer procesos
    col_indices = {h: i for i, h in enumerate(headers)}

    def get_val(row, field_name):
        col = mapping.get(field_name)
        if not col or col not in col_indices:
            return None
        idx = col_indices[col]
        val = row[idx] if idx < len(row) else None
        if val is None or str(val).strip() in ("", "None"):
            return None
        return val

    processes = []
    for row in raw_rows:
        name = get_val(row, "name")
        if not name:
            continue
        crit = (get_val(row, "criticality") or "medium").lower()
        if crit not in ("critical", "high", "medium", "low"):
            warnings.append(f"Proceso '{name}': criticidad '{crit}' invalida, se usara 'medium'")
            crit = "medium"
        processes.append({
            "name": str(name).strip(),
            "criticality": crit,
            "rto_hours": _safe_int(get_val(row, "rto_hours")),
            "rpo_hours": _safe_int(get_val(row, "rpo_hours")),
            "mtpd_hours": _safe_int(get_val(row, "mtpd_hours")),
            "description": str(get_val(row, "description")).strip()
                if get_val(row, "description") else None,
            "priority": _safe_int(get_val(row, "priority")),
            "activation_criteria": str(get_val(row, "activation_criteria")).strip()
                if get_val(row, "activation_criteria") else None,
            "owner_email": str(get_val(row, "owner_email")).strip()
                if get_val(row, "owner_email") else None,
        })

    errors = []
    if missing:
        errors.append(f"Campos obligatorios no encontrados: {', '.join(missing)}")

    return {
        "errors": errors,
        "summary": {
            "processes_found": len(processes),
            "dependencies_found": 0,
            "suppliers_found": 0,
        },
        "processes": processes[:3],
        "all_processes": processes,
        "all_dependencies": [],
        "all_suppliers": [],
        "ai_mapping": mapping,
        "ai_warnings": warnings,
        "ai_confidence": ai_result.get("confidence", 0.5),
    }
